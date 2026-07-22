# Implementa: RF06, RF07, RF08 (UC06, UC07) — ver docs/relatorio/04_analise_desenvolvimento/
#
# Qualificação docente e grade curricular são aditivas e idempotentes por par
# (ao contrário do diálogo manual, que substitui o conjunto todo) — ver
# app/services/importacao_service.py.
from io import BytesIO

import openpyxl
from sqlmodel import Session, SQLModel, create_engine, select

import app.models  # noqa: F401 - garante que todos os modelos entram no metadata
from app.models.curso import Curso
from app.models.disciplina import Disciplina
from app.models.plano_curricular import PlanoCurricular
from app.models.plano_curricular_disciplina import PlanoCurricularDisciplina
from app.models.professor import Professor
from app.models.professor_disciplina import ProfessorDisciplina
from app.models.turma import Turma
from app.services import importacao_service


def _criar_engine_teste():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False})
    SQLModel.metadata.create_all(engine)
    return engine


def _planilha(cabecalho: list[str], linhas: list[tuple]) -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(cabecalho)
    for linha in linhas:
        sheet.append(linha)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _planilha_multifolha(folhas: dict[str, tuple[list[str], list[tuple]]]) -> bytes:
    """Um único .xlsx com uma folha por entidade — usado para testar o caso
    'plano curricular + grade curricular no mesmo ficheiro, folhas diferentes'."""
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for nome_folha, (cabecalho, linhas) in folhas.items():
        sheet = workbook.create_sheet(title=nome_folha)
        sheet.append(cabecalho)
        for linha in linhas:
            sheet.append(linha)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_importar_qualificacoes_associa_professor_a_disciplina():
    engine = _criar_engine_teste()
    with Session(engine) as session:
        professor = Professor(nome="Ana Costa", email="ana@isaf.co.ao", classificacao=4, vinculo_casa=False)
        disciplina = Disciplina(codigo="POO", nome="Programação Orientada a Objectos")
        session.add(professor)
        session.add(disciplina)
        session.commit()
        session.refresh(professor)
        session.refresh(disciplina)

        conteudo = _planilha(["professor_email", "disciplina_codigo"], [("ana@isaf.co.ao", "POO")])
        relatorio = importacao_service.importar("qualificacoes", conteudo, session)

        assert relatorio.importados == 1
        assert relatorio.erros == []
        pares = list(session.exec(select(ProfessorDisciplina)))
        assert len(pares) == 1
        assert pares[0].professor_id == professor.id
        assert pares[0].disciplina_id == disciplina.id


def test_importar_qualificacoes_e_idempotente_e_reporta_email_inexistente():
    engine = _criar_engine_teste()
    with Session(engine) as session:
        professor = Professor(nome="Ana Costa", email="ana@isaf.co.ao", classificacao=4, vinculo_casa=False)
        disciplina = Disciplina(codigo="POO", nome="Programação Orientada a Objectos")
        session.add(professor)
        session.add(disciplina)
        session.commit()

        conteudo = _planilha(
            ["professor_email", "disciplina_codigo"],
            [("ana@isaf.co.ao", "POO"), ("ana@isaf.co.ao", "POO"), ("inexistente@isaf.co.ao", "POO")],
        )
        relatorio = importacao_service.importar("qualificacoes", conteudo, session)

        assert relatorio.importados == 1
        assert relatorio.ignorados_idempotencia == 1
        assert len(relatorio.erros) == 1
        assert relatorio.erros[0].campo == "professor_email"


def test_importar_planos_curriculares_e_idempotente():
    engine = _criar_engine_teste()
    with Session(engine) as session:
        curso = Curso(codigo="CC", nome="Ciência da Computação")
        session.add(curso)
        session.commit()

        conteudo = _planilha(["curso_codigo", "ano", "semestre"], [("CC", 1, "1")])
        relatorio = importacao_service.importar("planos_curriculares", conteudo, session)
        assert relatorio.importados == 1
        assert relatorio.erros == []

        relatorio2 = importacao_service.importar("planos_curriculares", conteudo, session)
        assert relatorio2.importados == 0
        assert relatorio2.ignorados_idempotencia == 1


def test_importar_grade_curricular_associa_plano_curricular_a_disciplina_com_carga_horaria():
    engine = _criar_engine_teste()
    with Session(engine) as session:
        curso = Curso(codigo="CC", nome="Ciência da Computação")
        session.add(curso)
        session.commit()
        session.refresh(curso)
        plano = PlanoCurricular(curso_id=curso.id, ano=1, semestre="1")
        disciplina = Disciplina(codigo="ALG201", nome="Algoritmos")
        session.add(plano)
        session.add(disciplina)
        session.commit()

        conteudo = _planilha(
            ["curso_codigo", "ano", "semestre", "disciplina_codigo", "carga_horaria_semanal"],
            [("CC", 1, "1", "ALG201", 4)],
        )
        relatorio = importacao_service.importar("grade_curricular", conteudo, session)

        assert relatorio.importados == 1
        assert relatorio.erros == []

        # Reimportar a mesma linha é ignorado por idempotência (não duplica, não atualiza).
        relatorio2 = importacao_service.importar("grade_curricular", conteudo, session)
        assert relatorio2.importados == 0
        assert relatorio2.ignorados_idempotencia == 1


def test_importar_planos_curriculares_com_folha_de_grade_curricular_no_mesmo_ficheiro():
    """RF06 — o Gestor pode preparar um único .xlsx com a folha 'planos_curriculares' e a
    folha 'grade_curricular' lado a lado; ambas são importadas na mesma chamada."""
    engine = _criar_engine_teste()
    with Session(engine) as session:
        curso = Curso(codigo="CC", nome="Ciência da Computação")
        disciplina = Disciplina(codigo="ALG201", nome="Algoritmos")
        session.add(curso)
        session.add(disciplina)
        session.commit()

        conteudo = _planilha_multifolha(
            {
                "planos_curriculares": (
                    ["curso_codigo", "ano", "semestre"],
                    [("CC", 1, "1")],
                ),
                "grade_curricular": (
                    ["curso_codigo", "ano", "semestre", "disciplina_codigo", "carga_horaria_semanal"],
                    [("CC", 1, "1", "ALG201", 4)],
                ),
            }
        )

        relatorio = importacao_service.importar("planos_curriculares", conteudo, session)

        assert relatorio.importados == 2  # 1 plano curricular + 1 item de grade curricular
        assert relatorio.erros == []

        plano = session.exec(
            select(PlanoCurricular).where(PlanoCurricular.curso_id == curso.id, PlanoCurricular.ano == 1)
        ).first()
        assert plano is not None
        grade = list(
            session.exec(
                select(PlanoCurricularDisciplina).where(PlanoCurricularDisciplina.plano_curricular_id == plano.id)
            )
        )
        assert len(grade) == 1
        assert grade[0].carga_horaria_semanal == 4


def test_importar_turmas_exige_plano_curricular_existente():
    """turmas referencia um PlanoCurricular por curso_codigo+ano+semestre — tem de já existir."""
    engine = _criar_engine_teste()
    with Session(engine) as session:
        curso = Curso(codigo="CC", nome="Ciência da Computação")
        session.add(curso)
        session.commit()

        conteudo = _planilha(
            ["codigo", "nome", "ano_letivo", "turno", "numero_alunos", "curso_codigo", "ano", "semestre"],
            [("CC-1A", "Turma 1A", 2026, "manha", 30, "CC", 1, "1")],
        )
        relatorio = importacao_service.importar("turmas", conteudo, session)

        assert relatorio.importados == 0
        assert len(relatorio.erros) == 1
        assert relatorio.erros[0].campo == "curso_codigo/ano/semestre"


def test_importar_turmas_com_plano_curricular_existente():
    engine = _criar_engine_teste()
    with Session(engine) as session:
        curso = Curso(codigo="CC", nome="Ciência da Computação")
        session.add(curso)
        session.commit()
        session.refresh(curso)
        plano = PlanoCurricular(curso_id=curso.id, ano=1, semestre="1")
        session.add(plano)
        session.commit()

        conteudo = _planilha(
            ["codigo", "nome", "ano_letivo", "turno", "numero_alunos", "curso_codigo", "ano", "semestre"],
            [("CC-1A", "Turma 1A", 2026, "manha", 30, "CC", 1, "1")],
        )
        relatorio = importacao_service.importar("turmas", conteudo, session)

        assert relatorio.importados == 1
        assert relatorio.erros == []

        turma = session.exec(select(Turma).where(Turma.codigo == "CC-1A")).first()
        assert turma is not None
        assert turma.plano_curricular_id == plano.id


def test_importar_grade_curricular_reporta_plano_curricular_ou_disciplina_inexistente():
    engine = _criar_engine_teste()
    with Session(engine) as session:
        conteudo = _planilha(
            ["curso_codigo", "ano", "semestre", "disciplina_codigo", "carga_horaria_semanal"],
            [("NAO-EXISTE", 1, "1", "NAO-EXISTE", 4)],
        )
        relatorio = importacao_service.importar("grade_curricular", conteudo, session)

        assert relatorio.importados == 0
        assert len(relatorio.erros) == 1
        assert relatorio.erros[0].campo == "curso_codigo"
