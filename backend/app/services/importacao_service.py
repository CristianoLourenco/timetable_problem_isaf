# Implementa: RF06, RF07, RF08 (UC06, UC07) — ver docs/04_04_analise_desenvolvimento.md
# Fluxo (ver skill excel-import-openpyxl): parse -> validar (RF07) -> gravar com idempotência por
# chave única (RF08). Nunca aborta no primeiro erro de linha — reporta todos.
from collections.abc import Callable, Sequence
from io import BytesIO
from typing import Any

import openpyxl
from sqlmodel import Session

from app.core.calendario import TurnoEnum
from app.models.curso import Curso
from app.models.disciplina import Disciplina
from app.models.plano_curricular import PlanoCurricular
from app.models.professor import Professor
from app.models.sala import Sala
from app.models.turma import Turma
from app.repositories.curso_repository import CursoRepository
from app.repositories.disciplina_repository import DisciplinaRepository
from app.repositories.plano_curricular_disciplina_repository import PlanoCurricularDisciplinaRepository
from app.repositories.plano_curricular_repository import PlanoCurricularRepository
from app.repositories.professor_disciplina_repository import ProfessorDisciplinaRepository
from app.repositories.professor_repository import ProfessorRepository
from app.repositories.sala_repository import SalaRepository
from app.repositories.turma_repository import TurmaRepository
from app.schemas.importacao_schema import ErroImportacaoSchema, RelatorioImportacaoSchema

COLUNAS_ESPERADAS: dict[str, list[str]] = {
    "cursos": ["codigo", "nome"],
    "professores": ["nome", "email", "classificacao", "vinculo_casa"],
    "disciplinas": ["codigo", "nome"],
    "salas": ["codigo", "nome", "capacidade"],
    # planos_curriculares depende de "cursos" já estar importado (curso_codigo tem de existir)
    "planos_curriculares": ["curso_codigo", "ano", "semestre"],
    # turmas depende de "planos_curriculares" já estar importado
    "turmas": ["codigo", "nome", "ano_letivo", "turno", "numero_alunos", "curso_codigo", "ano", "semestre"],
    # qualificacoes depende de professores/disciplinas; grade_curricular depende de
    # planos_curriculares/disciplinas já importados
    "qualificacoes": ["professor_email", "disciplina_codigo"],
    "grade_curricular": ["curso_codigo", "ano", "semestre", "disciplina_codigo", "carga_horaria_semanal"],
}

# Um PlanoCurricular sem grade curricular (e um Professor sem qualificação) não serve
# para nada no solver — por isso o mesmo ficheiro .xlsx pode trazer a entidade principal
# e o seu complemento em folhas (sheets) diferentes, e ambas são importadas na mesma
# chamada. A folha do complemento só é processada se existir (identificada pelo nome,
# não pela posição) — um ficheiro de "planos_curriculares" sem folha de grade curricular
# continua a funcionar como antes.
COMPLEMENTOS: dict[str, str] = {
    "planos_curriculares": "grade_curricular",
    "professores": "qualificacoes",
}


_VALORES_VERDADEIROS_PT = {"sim", "true", "1", "x", "verdadeiro"}


def _parse_booleano_pt(valor: Any) -> bool:
    """`bool(valor)` nativo trataria qualquer string não vazia (incl. "Não") como
    True — uma célula Excel com "Não" inverteria silenciosamente a intenção do
    Gestor. Só os valores desta lista (case/espaço-insensitive) contam como True."""
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, int | float):
        return bool(valor)
    if isinstance(valor, str):
        return valor.strip().lower() in _VALORES_VERDADEIROS_PT
    return False


def _abrir_workbook(conteudo: bytes):
    try:
        return openpyxl.load_workbook(BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as exc:  # ficheiro corrompido ou não é .xlsx
        raise ValueError("Ficheiro Excel inválido ou corrompido.") from exc


def _obter_folha_por_nome(workbook, nome: str):
    """Procura uma folha pelo nome (case-insensitive) — nunca pela posição, para que a
    ordem das folhas no ficheiro seja irrelevante."""
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().lower() == nome:
            return workbook[sheet_name]
    return None


def _ler_folha(sheet) -> tuple[list[str], list[tuple[Any, ...]]]:
    linhas = list(sheet.iter_rows(values_only=True))
    if not linhas:
        return [], []
    cabecalho = [str(c).strip().lower() if c is not None else "" for c in linhas[0]]
    return cabecalho, linhas[1:]


def _processar_folha(
    entidade: str, sheet, session: Session, relatorio: RelatorioImportacaoSchema
) -> None:
    cabecalho, linhas_dados = _ler_folha(sheet)
    relatorio.total_linhas += len(linhas_dados)

    esperadas = COLUNAS_ESPERADAS[entidade]
    if cabecalho[: len(esperadas)] != esperadas:
        relatorio.erros.append(
            ErroImportacaoSchema(
                linha=1,
                campo="cabecalho",
                motivo=f"[{entidade}] Colunas esperadas (nesta ordem): {', '.join(esperadas)}",
            )
        )
        return

    _IMPORTADORES[entidade](session, cabecalho, linhas_dados, relatorio)


def _importar_professores(
    session: Session, colunas: list[str], linhas: Sequence[tuple[Any, ...]], relatorio: RelatorioImportacaoSchema
) -> None:
    repo = ProfessorRepository(session)
    for numero_linha, linha in enumerate(linhas, start=2):
        dados = dict(zip(colunas, linha, strict=False))
        nome, email = dados.get("nome"), dados.get("email")
        if not nome:
            relatorio.erros.append(ErroImportacaoSchema(linha=numero_linha, campo="nome", motivo="Nome em falta"))
            continue
        if not email:
            relatorio.erros.append(
                ErroImportacaoSchema(linha=numero_linha, campo="email", motivo="Email institucional em falta")
            )
            continue
        classificacao = int(dados.get("classificacao") or 3)
        if not (1 <= classificacao <= 5):
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha, campo="classificacao", motivo="Classificação deve ser entre 1 e 5"
                )
            )
            continue
        if repo.get_by_email(str(email)):
            relatorio.ignorados_idempotencia += 1
            continue
        repo.create(
            Professor(
                nome=str(nome),
                email=str(email),
                classificacao=classificacao,
                vinculo_casa=_parse_booleano_pt(dados.get("vinculo_casa")),
            )
        )
        relatorio.importados += 1


def _importar_cursos(
    session: Session, colunas: list[str], linhas: Sequence[tuple[Any, ...]], relatorio: RelatorioImportacaoSchema
) -> None:
    repo = CursoRepository(session)
    for numero_linha, linha in enumerate(linhas, start=2):
        dados = dict(zip(colunas, linha, strict=False))
        codigo, nome = dados.get("codigo"), dados.get("nome")
        if not codigo or not nome:
            relatorio.erros.append(
                ErroImportacaoSchema(linha=numero_linha, campo="codigo/nome", motivo="Código ou nome em falta")
            )
            continue
        if repo.get_by_codigo(str(codigo)):
            relatorio.ignorados_idempotencia += 1
            continue
        repo.create(Curso(codigo=str(codigo), nome=str(nome)))
        relatorio.importados += 1


def _importar_planos_curriculares(
    session: Session, colunas: list[str], linhas: Sequence[tuple[Any, ...]], relatorio: RelatorioImportacaoSchema
) -> None:
    plano_repo = PlanoCurricularRepository(session)
    curso_repo = CursoRepository(session)
    for numero_linha, linha in enumerate(linhas, start=2):
        dados = dict(zip(colunas, linha, strict=False))
        curso_codigo, ano_bruto, semestre = dados.get("curso_codigo"), dados.get("ano"), dados.get("semestre")
        if not curso_codigo or not ano_bruto or not semestre:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha,
                    campo="curso_codigo/ano/semestre",
                    motivo="Curso, ano ou semestre em falta",
                )
            )
            continue
        curso = curso_repo.get_by_codigo(str(curso_codigo))
        if curso is None:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha, campo="curso_codigo", motivo=f"Curso '{curso_codigo}' não existe"
                )
            )
            continue
        ano = int(ano_bruto)
        if plano_repo.get_by_curso_ano_semestre(curso.id, ano, str(semestre)):  # type: ignore[arg-type]
            relatorio.ignorados_idempotencia += 1
            continue
        plano_repo.create(PlanoCurricular(curso_id=curso.id, ano=ano, semestre=str(semestre)))  # type: ignore[arg-type]
        relatorio.importados += 1


def _importar_disciplinas(
    session: Session, colunas: list[str], linhas: Sequence[tuple[Any, ...]], relatorio: RelatorioImportacaoSchema
) -> None:
    repo = DisciplinaRepository(session)
    for numero_linha, linha in enumerate(linhas, start=2):
        dados = dict(zip(colunas, linha, strict=False))
        codigo, nome = dados.get("codigo"), dados.get("nome")
        if not codigo or not nome:
            relatorio.erros.append(
                ErroImportacaoSchema(linha=numero_linha, campo="codigo/nome", motivo="Código ou nome em falta")
            )
            continue
        if repo.get_by_codigo(str(codigo)):
            relatorio.ignorados_idempotencia += 1
            continue
        repo.create(Disciplina(codigo=str(codigo), nome=str(nome)))
        relatorio.importados += 1


def _importar_salas(
    session: Session, colunas: list[str], linhas: Sequence[tuple[Any, ...]], relatorio: RelatorioImportacaoSchema
) -> None:
    repo = SalaRepository(session)
    for numero_linha, linha in enumerate(linhas, start=2):
        dados = dict(zip(colunas, linha, strict=False))
        codigo, nome, capacidade = dados.get("codigo"), dados.get("nome"), dados.get("capacidade")
        if not codigo or not nome:
            relatorio.erros.append(
                ErroImportacaoSchema(linha=numero_linha, campo="codigo/nome", motivo="Código ou nome em falta")
            )
            continue
        if not capacidade or int(capacidade) <= 0:
            relatorio.erros.append(
                ErroImportacaoSchema(linha=numero_linha, campo="capacidade", motivo="Capacidade deve ser maior que 0")
            )
            continue
        if repo.get_by_codigo(str(codigo)):
            relatorio.ignorados_idempotencia += 1
            continue
        repo.create(Sala(codigo=str(codigo), nome=str(nome), capacidade=int(capacidade)))
        relatorio.importados += 1


def _importar_turmas(
    session: Session, colunas: list[str], linhas: Sequence[tuple[Any, ...]], relatorio: RelatorioImportacaoSchema
) -> None:
    turma_repo = TurmaRepository(session)
    curso_repo = CursoRepository(session)
    plano_repo = PlanoCurricularRepository(session)
    for numero_linha, linha in enumerate(linhas, start=2):
        dados = dict(zip(colunas, linha, strict=False))
        codigo = dados.get("codigo")
        curso_codigo, ano_bruto, semestre = dados.get("curso_codigo"), dados.get("ano"), dados.get("semestre")
        if not codigo:
            relatorio.erros.append(ErroImportacaoSchema(linha=numero_linha, campo="codigo", motivo="Código em falta"))
            continue
        if not curso_codigo or not ano_bruto or not semestre:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha,
                    campo="curso_codigo/ano/semestre",
                    motivo="Curso, ano ou semestre do plano curricular em falta",
                )
            )
            continue
        curso = curso_repo.get_by_codigo(str(curso_codigo))
        if curso is None:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha, campo="curso_codigo", motivo=f"Curso '{curso_codigo}' não existe"
                )
            )
            continue
        plano = plano_repo.get_by_curso_ano_semestre(curso.id, int(ano_bruto), str(semestre))  # type: ignore[arg-type]
        if plano is None:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha,
                    campo="curso_codigo/ano/semestre",
                    motivo=(
                        f"Não existe PlanoCurricular para curso '{curso_codigo}', "
                        f"ano {ano_bruto}, semestre '{semestre}' — importar planos_curriculares primeiro"
                    ),
                )
            )
            continue
        numero_alunos = int(dados.get("numero_alunos") or 0)
        if numero_alunos <= 0:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha, campo="numero_alunos", motivo="Número de alunos deve ser maior que 0"
                )
            )
            continue
        turno_bruto = str(dados.get("turno") or "").strip().lower()
        try:
            turno = TurnoEnum(turno_bruto)
        except ValueError:
            valores_validos = ", ".join(t.value for t in TurnoEnum)
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha,
                    campo="turno",
                    motivo=f"Turno '{turno_bruto}' inválido — valores aceites: {valores_validos}",
                )
            )
            continue
        if turma_repo.get_by_codigo(str(codigo)):
            relatorio.ignorados_idempotencia += 1
            continue
        turma_repo.create(
            Turma(
                codigo=str(codigo),
                nome=str(dados.get("nome") or ""),
                ano_letivo=int(dados.get("ano_letivo") or 0),
                turno=turno,
                numero_alunos=numero_alunos,
                plano_curricular_id=plano.id,  # type: ignore[arg-type]
            )
        )
        relatorio.importados += 1


def _importar_qualificacoes(
    session: Session, colunas: list[str], linhas: Sequence[tuple[Any, ...]], relatorio: RelatorioImportacaoSchema
) -> None:
    """Qualificação docente (professor↔disciplinas) — RN filtro obrigatório do solver.
    Aditiva: cada linha é um par a acrescentar, nunca substitui os pares já existentes
    (isso é o que o diálogo manual `POST /professores/{id}/disciplinas` faz)."""
    qualificacao_repo = ProfessorDisciplinaRepository(session)
    professor_repo = ProfessorRepository(session)
    disciplina_repo = DisciplinaRepository(session)
    for numero_linha, linha in enumerate(linhas, start=2):
        dados = dict(zip(colunas, linha, strict=False))
        email, disciplina_codigo = dados.get("professor_email"), dados.get("disciplina_codigo")
        if not email or not disciplina_codigo:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha,
                    campo="professor_email/disciplina_codigo",
                    motivo="Email do professor ou código da disciplina em falta",
                )
            )
            continue
        professor = professor_repo.get_by_email(str(email))
        if professor is None:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha, campo="professor_email", motivo=f"Professor '{email}' não existe"
                )
            )
            continue
        disciplina = disciplina_repo.get_by_codigo(str(disciplina_codigo))
        if disciplina is None:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha,
                    campo="disciplina_codigo",
                    motivo=f"Disciplina '{disciplina_codigo}' não existe",
                )
            )
            continue
        if qualificacao_repo.existe(professor.id, disciplina.id):  # type: ignore[arg-type]
            relatorio.ignorados_idempotencia += 1
            continue
        qualificacao_repo.adicionar(professor.id, disciplina.id)  # type: ignore[arg-type]
        relatorio.importados += 1


def _importar_grade_curricular(
    session: Session, colunas: list[str], linhas: Sequence[tuple[Any, ...]], relatorio: RelatorioImportacaoSchema
) -> None:
    """Grade curricular (PlanoCurricular↔disciplina+carga horária) — pré-requisito de
    dados do solver. O plano é identificado por curso+ano+semestre (chave natural, mais
    prática para o Gestor preencher do que um id interno). Aditiva, mesmo princípio de
    `_importar_qualificacoes`."""
    itens_repo = PlanoCurricularDisciplinaRepository(session)
    plano_repo = PlanoCurricularRepository(session)
    curso_repo = CursoRepository(session)
    disciplina_repo = DisciplinaRepository(session)
    for numero_linha, linha in enumerate(linhas, start=2):
        dados = dict(zip(colunas, linha, strict=False))
        curso_codigo, ano_bruto, semestre = dados.get("curso_codigo"), dados.get("ano"), dados.get("semestre")
        disciplina_codigo = dados.get("disciplina_codigo")
        if not curso_codigo or not ano_bruto or not semestre or not disciplina_codigo:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha,
                    campo="curso_codigo/ano/semestre/disciplina_codigo",
                    motivo="Curso, ano, semestre ou disciplina em falta",
                )
            )
            continue
        curso = curso_repo.get_by_codigo(str(curso_codigo))
        if curso is None:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha, campo="curso_codigo", motivo=f"Curso '{curso_codigo}' não existe"
                )
            )
            continue
        plano = plano_repo.get_by_curso_ano_semestre(curso.id, int(ano_bruto), str(semestre))  # type: ignore[arg-type]
        if plano is None:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha,
                    campo="curso_codigo/ano/semestre",
                    motivo=(
                        f"Não existe PlanoCurricular para curso '{curso_codigo}', "
                        f"ano {ano_bruto}, semestre '{semestre}'"
                    ),
                )
            )
            continue
        disciplina = disciplina_repo.get_by_codigo(str(disciplina_codigo))
        if disciplina is None:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha,
                    campo="disciplina_codigo",
                    motivo=f"Disciplina '{disciplina_codigo}' não existe",
                )
            )
            continue
        carga_horaria_semanal = int(dados.get("carga_horaria_semanal") or 0)
        if carga_horaria_semanal <= 0:
            relatorio.erros.append(
                ErroImportacaoSchema(
                    linha=numero_linha,
                    campo="carga_horaria_semanal",
                    motivo="Carga horária semanal deve ser maior que 0",
                )
            )
            continue
        if itens_repo.existe(plano.id, disciplina.id):  # type: ignore[arg-type]
            relatorio.ignorados_idempotencia += 1
            continue
        itens_repo.adicionar(plano.id, disciplina.id, carga_horaria_semanal)  # type: ignore[arg-type]
        relatorio.importados += 1


_IMPORTADORES: dict[str, Callable[[Session, list[str], Sequence[tuple[Any, ...]], RelatorioImportacaoSchema], None]] = {
    "cursos": _importar_cursos,
    "planos_curriculares": _importar_planos_curriculares,
    "professores": _importar_professores,
    "disciplinas": _importar_disciplinas,
    "salas": _importar_salas,
    "turmas": _importar_turmas,
    "qualificacoes": _importar_qualificacoes,
    "grade_curricular": _importar_grade_curricular,
}


def importar(entidade: str, conteudo: bytes, session: Session) -> RelatorioImportacaoSchema:
    if entidade not in COLUNAS_ESPERADAS:
        raise ValueError(f"Entidade '{entidade}' não suportada para importação.")

    workbook = _abrir_workbook(conteudo)
    relatorio = RelatorioImportacaoSchema()

    # Folha principal: procurada pelo nome da entidade; se o ficheiro só tiver uma
    # folha sem esse nome (caso comum de upload de uma única entidade), usa-se a
    # folha ativa — mantém compatível com ficheiros de uma folha só.
    folha_principal = _obter_folha_por_nome(workbook, entidade) or workbook.active
    _processar_folha(entidade, folha_principal, session, relatorio)

    complemento = COMPLEMENTOS.get(entidade)
    if complemento:
        folha_complemento = _obter_folha_por_nome(workbook, complemento)
        if folha_complemento is not None:
            _processar_folha(complemento, folha_complemento, session, relatorio)

    return relatorio
